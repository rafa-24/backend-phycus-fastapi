from decimal import Decimal, InvalidOperation
from io import BytesIO

from openpyxl import Workbook, load_workbook

EXCEL_HEADERS = [
    "categoria",
    "nombre",
    "descripcion",
    "precio",
    "imagen_url",
    "activo",
]


def create_products_workbook(rows: list[dict]) -> BytesIO:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Productos"
    worksheet.append(EXCEL_HEADERS)

    for row in rows:
        worksheet.append(
            [
                row.get("category_name", ""),
                row["name"],
                row.get("description") or "",
                float(row["price"]),
                row.get("image_url") or "",
                "si" if row.get("is_active", True) else "no",
            ]
        )

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def _normalize_header(value) -> str:
    if value is None:
        return ""

    return str(value).strip().lower()


def _parse_bool(value) -> bool:
    if value is None or str(value).strip() == "":
        return True

    normalized = str(value).strip().lower()

    if normalized in {"si", "sí", "true", "1", "activo", "yes"}:
        return True

    if normalized in {"no", "false", "0", "inactivo"}:
        return False

    raise ValueError("El valor de 'activo' debe ser 'si' o 'no'.")


def _parse_price(value) -> Decimal:
    if value is None or str(value).strip() == "":
        raise ValueError("El precio es obligatorio.")

    try:
        price = Decimal(str(value).strip().replace(",", "."))
    except (InvalidOperation, ValueError) as exc:
        raise ValueError("El precio no tiene un formato válido.") from exc

    if price <= 0:
        raise ValueError("El precio debe ser mayor a cero.")

    return price


def parse_products_workbook(file_content: bytes) -> dict:
    workbook = load_workbook(BytesIO(file_content), read_only=True, data_only=True)
    worksheet = workbook.active

    rows = worksheet.iter_rows(values_only=True)
    header_row = next(rows, None)

    if not header_row:
        raise ValueError("El archivo Excel está vacío.")

    headers = [_normalize_header(value) for value in header_row]
    header_index = {header: index for index, header in enumerate(headers) if header}

    required_headers = {"categoria", "nombre", "precio"}
    missing_headers = required_headers - set(header_index)

    if missing_headers:
        missing = ", ".join(sorted(missing_headers))
        raise ValueError(f"Faltan columnas obligatorias en el Excel: {missing}.")

    valid_rows = []
    skipped_rows = []

    for row_number, row in enumerate(rows, start=2):
        if not row or all(cell is None or str(cell).strip() == "" for cell in row):
            continue

        def get_value(column_name: str):
            index = header_index.get(column_name)
            if index is None or index >= len(row):
                return None
            return row[index]

        category_name = get_value("categoria")
        name = get_value("nombre")
        description = get_value("descripcion")
        price_value = get_value("precio")
        image_url = get_value("imagen_url")
        is_active_value = get_value("activo")

        category_name = str(category_name).strip() if category_name is not None else ""
        name = str(name).strip() if name is not None else ""

        if not category_name:
            skipped_rows.append(
                {
                    "row_number": row_number,
                    "reason": "La categoría está vacía.",
                    "categoria": category_name or None,
                    "nombre": name or None,
                }
            )
            continue

        if not name:
            skipped_rows.append(
                {
                    "row_number": row_number,
                    "reason": "El nombre está vacío.",
                    "categoria": category_name,
                    "nombre": None,
                }
            )
            continue

        try:
            price = _parse_price(price_value)
        except ValueError as exc:
            skipped_rows.append(
                {
                    "row_number": row_number,
                    "reason": str(exc),
                    "categoria": category_name,
                    "nombre": name,
                }
            )
            continue

        try:
            is_active = _parse_bool(is_active_value)
        except ValueError as exc:
            skipped_rows.append(
                {
                    "row_number": row_number,
                    "reason": str(exc),
                    "categoria": category_name,
                    "nombre": name,
                }
            )
            continue

        valid_rows.append(
            {
                "row_number": row_number,
                "category_name": category_name,
                "name": name,
                "description": str(description).strip()
                if description is not None and str(description).strip()
                else None,
                "price": price,
                "image_url": str(image_url).strip()
                if image_url is not None and str(image_url).strip()
                else None,
                "is_active": is_active,
            }
        )

    workbook.close()

    if not valid_rows and not skipped_rows:
        raise ValueError("El archivo Excel no contiene productos para importar.")

    if not valid_rows:
        raise ValueError(
            "No fue posible importar ningún producto. Todas las filas tienen errores."
        )

    return {
        "valid_rows": valid_rows,
        "skipped_rows": skipped_rows,
    }
