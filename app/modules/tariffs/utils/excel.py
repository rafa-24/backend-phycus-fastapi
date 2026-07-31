from decimal import Decimal, InvalidOperation
from io import BytesIO
import re

from openpyxl import Workbook, load_workbook

EXCEL_HEADERS = [
    "localidad",
    "barrio",
    "tarifa",
    "tarifa_enrutar",
    "ciudad",
    "activo",
]


def create_tariffs_template() -> BytesIO:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Tarifas"
    worksheet.append(EXCEL_HEADERS)
    worksheet.append(
        ["RIO MAR", "Altamira", 6000, "", "Barranquilla", "si"]
    )
    worksheet.append(
        ["NORTE CENTRO HISTORICO", "El Prado", 7000, "", "Barranquilla", "si"]
    )

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def _normalize_header(value) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = (
        text.replace("á", "a")
        .replace("é", "e")
        .replace("í", "i")
        .replace("ó", "o")
        .replace("ú", "u")
        .replace("ñ", "n")
    )
    text = re.sub(r"\s+", "_", text)
    aliases = {
        "localidad": "localidad",
        "barrio": "barrio",
        "tarifa": "tarifa",
        "tarifa_enrutar": "tarifa_enrutar",
        "t_para_enrutar": "tarifa_enrutar",
        "t.para_enrutar": "tarifa_enrutar",
        "ciudad": "ciudad",
        "city": "ciudad",
        "activo": "activo",
    }
    return aliases.get(text, text)


def _parse_bool(value) -> bool:
    if value is None or str(value).strip() == "":
        return True
    normalized = str(value).strip().lower()
    if normalized in {"si", "sí", "true", "1", "activo", "yes"}:
        return True
    if normalized in {"no", "false", "0", "inactivo"}:
        return False
    raise ValueError("El valor de 'activo' debe ser 'si' o 'no'.")


def _parse_money(value) -> Decimal | None:
    if value is None or str(value).strip() == "":
        return None
    try:
        amount = Decimal(str(value).strip().replace(",", "."))
    except (InvalidOperation, ValueError) as exc:
        raise ValueError("El valor monetario no es válido.") from exc
    if amount < 0:
        raise ValueError("El valor monetario no puede ser negativo.")
    return amount


def _clean_text(value) -> str:
    if value is None:
        return ""
    text = str(value).replace("\xa0", " ").replace("\u00a0", " ")
    text = re.sub(r"^[^\wÁÉÍÓÚáéíóúÑñÜü]+", "", text)
    return re.sub(r"\s+", " ", text).strip()


def parse_tariffs_workbook(file_content: bytes) -> list[dict]:
    workbook = load_workbook(BytesIO(file_content), read_only=True, data_only=True)
    worksheet = workbook.active

    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError("El archivo Excel está vacío.")

    header_row_index = 0
    headers: list[str] = []
    for index, row in enumerate(rows[:15]):
        normalized = [_normalize_header(cell) for cell in row]
        if "localidad" in normalized and "barrio" in normalized:
            headers = normalized
            header_row_index = index
            break

    if not headers:
        raise ValueError(
            "No se encontraron las columnas obligatorias: localidad, barrio."
        )

    parsed_rows: list[dict] = []

    for offset, row in enumerate(rows[header_row_index + 1 :], start=header_row_index + 2):
        values = {
            headers[i]: row[i] if i < len(row) else None for i in range(len(headers))
        }

        localidad = _clean_text(values.get("localidad"))
        barrio = _clean_text(values.get("barrio"))

        # Excel estilo 411: localidad se rellena solo en la primera fila del grupo
        if not barrio and not localidad:
            continue

        try:
            tarifa = _parse_money(values.get("tarifa"))
            tarifa_enrutar = _parse_money(values.get("tarifa_enrutar"))
            is_active = _parse_bool(values.get("activo"))
            city = _clean_text(values.get("ciudad")) or None
        except ValueError as exc:
            parsed_rows.append(
                {
                    "row_number": offset,
                    "error": str(exc),
                    "localidad": localidad or None,
                    "barrio": barrio or None,
                }
            )
            continue

        if not barrio:
            parsed_rows.append(
                {
                    "row_number": offset,
                    "error": "El barrio es obligatorio.",
                    "localidad": localidad or None,
                    "barrio": None,
                }
            )
            continue

        parsed_rows.append(
            {
                "row_number": offset,
                "localidad": localidad,
                "barrio": barrio,
                "tarifa": tarifa,
                "tarifa_enrutar": tarifa_enrutar,
                "city": city,
                "is_active": is_active,
            }
        )

    return parsed_rows
